#-*-coding:utf-8-*-
#@Time      :2019/5/1 0001 20:22
#@Author    :蓝天下的风
#@Email     :394845369@qq.com
#@File      :do_excel.py
#@Software  :PyCharm Community Edition

import openpyxl
from Common import http_request


class DoExcel:

    #初始化excel的name,sheet的name

    def __init__(self,file_name,sheet_name):
        #处理异常
        try:
            self.file_name=file_name
            self.sheet_name = sheet_name
            self.workbook = openpyxl.load_workbook(file_name)
            self.sheet = self.workbook[sheet_name]
        except ZeroDivisionError as e:
            print(e)


    #从excel中获取测试用例的
    def get_cases(self):
        max_row = self.sheet.max_row #获取最大行数

        cases = []#列表，存放所有的测试用例
        for r in range(2,max_row+1):
            #方法一：原始写法
            # case = {}
            # case['case_id'] = self.sheet.cell(row = r,column=1)
            # case['title'] = self.sheet.cell(row=r, column=2)
            #方法二：实例化类，将参数存放在实例化中
            case = Case() #实例化Case类
            case.case_id = self.sheet.cell(row = r,column=1).value
            case.title = self.sheet.cell(row=r, column=2).value
            case.url = self.sheet.cell(row=r, column=3).value
            case.data = self.sheet.cell(row=r, column=4).value
            case.method = self.sheet.cell(row=r, column=5).value
            case.expected = self.sheet.cell(row=r, column=6).value
            case.sql = self.sheet.cell(row=r, column=9).value
            cases.append(case)
        #读取完excel后，一定要记得关闭
        self.workbook.close()
        return cases #返回case列表

    '''
        将返回结果写入excel中
    '''
    def write_result(self,row,actual,result):
        #获取到一个sheet
        sheet = self.workbook[self.sheet_name]
        #每一张表中数据收集起来，返回一个列表
        sheet.cell(row,7).value = actual    #将获取的值写入到第row行，第7列
        sheet.cell(row,8).value = result    #将获取的值写入到第row行，第8列
        self.workbook.save(filename=self.file_name)     #保存
        self.workbook.close()   #关闭excel

'''
    测试用例类，每个测试用例，实际上就是一个它的实例
'''
class Case:

    def __init__(self):
        self.case_id = None
        self.title=None
        self.url = None
        self.data = None
        self.method =None
        self.expected = None
        self.actual=None
        self.result=None
        self.sql=None

'''
if __name__ == '__main__':

    from Common import contants
    do_excel = DoExcel(contants.case_file,sheet_name='login')
    cases = do_excel.get_cases()
    http_request = http_request.HttpRequest()
    for case in cases:
        #查看case中的获取的值
        # print(case.case_id)
        # print(case.url)
        # print(case.data)
        #返回一个字典，将所有的值返回
        # print(case.__dict__)
        print("case入参data数据:{}".format(case.data))  #第一次读取出来，为Sting类型，需要转换数据类型为字典dict
        resp = http_request.request(case.method,case.url,case.data)
        # print(resp.status_code)
        # print(resp.text)#响应文本
        #返回字典
        resp_dict = resp.json()
        print("返回值resp_dict数据:{}".format(resp_dict))
        actual = resp.text

        if case.expected == actual: #判断期望结果是否与实际结果一致
            #调用do_excel中写入excel的方法，写入case中第case_id+1行，第Actual列，通过，pass
            do_excel.write_result(case.case_id + 1,actual,'PASS')
            print("第{}个测试用例，测试通过！{}".format(case.case_id,'PASS'))
        else:
            #调用do_excel中写入excel的方法，写入case中第case_id+1行，第Actual列，不通过，Faile
            do_excel.write_result(case.case_id + 1,actual,'FAILE')
            print("第{}个测试用例，测试不通过！{}".format(case.case_id,'FAILE'))

'''
